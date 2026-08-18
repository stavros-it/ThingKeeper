"""Exporters: compressed JSON archive (.tkz), CSV, Excel (.xlsx), PDF report."""

from __future__ import annotations

import csv
import gzip
import json
import zipfile
from datetime import datetime
from pathlib import Path

from . import config
from .repository import (
    Item,
    all_items,
    counts_by,
    list_contacts,
    list_images,
    list_loans,
    total_depreciated_value,
    total_quantity,
    total_value,
    warranty_expired,
    warranty_expiring,
)


def _fmt(value) -> str:
    if value is None:
        return ""
    return str(value)


def export_csv(path: str | Path, items: list[Item] | None = None) -> Path:
    """Export items to a CSV file."""
    items = items if items is not None else all_items()
    path = Path(path)
    fieldnames = [
        "group_name", "type", "brand", "model", "info", "serial", "store",
        "purchase_date", "status", "quantity", "location", "warranty_end",
        "unit_price", "depreciation_years", "image_path",
    ]
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for it in items:
            row = it.as_dict()
            row.pop("id", None)
            writer.writerow({k: _fmt(row.get(k, "")) for k in fieldnames})
    return path


def export_archive(path: str | Path, items: list[Item] | None = None) -> Path:
    """Export items + attachments + loans + contacts to a .tkz zip archive.

    Layout inside the archive:
        items.json.gz        — gzipped JSON list of item dicts
        loans.json.gz        — gzipped JSON list of loan dicts
        contacts.json.gz     — gzipped JSON list of contact dicts
        attachments/<file>   — referenced image files
    """
    items = items if items is not None else all_items()
    path = Path(path)
    records = []
    attach_seen: set[str] = set()
    for it in items:
        rec = it.as_dict()
        if it.image_path:
            base = Path(it.image_path).name
            rec["image_path"] = base
            p = Path(it.image_path)
            if p.exists():
                attach_seen.add(str(p))
        # Include extra images from the item_images table.
        extra = []
        if it.id is not None:
            for _img_id, img_path in list_images(it.id):
                p = Path(img_path)
                extra.append(p.name)
                if p.exists():
                    attach_seen.add(str(p))
        rec["extra_images"] = extra
        records.append(rec)

    # Build a quick item-id -> serial lookup so loans can be re-mapped on import.
    items_by_id = {it.id: it for it in items if it.id is not None}
    loan_records = []
    for ln in list_loans():
        rec = ln.as_dict()
        item = items_by_id.get(ln.item_id)
        if item is not None:
            rec["_serial"] = item.serial
        loan_records.append(rec)
    contacts = [c.as_dict() for c in list_contacts()]

    payload = json.dumps(
        {"items": records, "loans": loan_records, "contacts": contacts},
        ensure_ascii=False,
        indent=2,
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("items.json.gz", gzip.compress(payload.encode("utf-8")))
        for ap in sorted(attach_seen):
            base = Path(ap).name
            zf.write(ap, arcname=f"attachments/{base}")
    return path


def export_excel(path: str | Path, items: list[Item] | None = None) -> Path:
    """Export items to an .xlsx workbook matching the original column layout."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    items = items if items is not None else all_items()
    path = Path(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "ID", "GROUP", "TYPE", "BRAND", "MODEL", "INFO", "PURCHASE",
        "SERIAL", "STORE", "STATUS", "QUANTITY", "LOCATION",
        "WARRANTY_END", "UNIT_PRICE", "DEPRECIATION_YEARS",
        "CREATED_AT", "UPDATED_AT",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions["E"].width = 40  # INFO
    ws.freeze_panes = "A2"

    for it in items:
        ws.append([
            it.id if it.id is not None else "",
            it.group_name, it.type, it.brand, it.model, it.info,
            it.purchase_date, it.serial, it.store, it.status,
            it.quantity, it.location, it.warranty_end,
            it.unit_price, it.depreciation_years,
            it.created_at, it.updated_at,
        ])

    wb.save(path)
    return path


def export_html(path: str | Path, items: list[Item] | None = None) -> Path:
    """Export items to a standalone HTML file for easy sharing / printing."""
    from html import escape

    items = items if items is not None else all_items()
    path = Path(path)
    headers = [
        "ID", "Group", "Type", "Brand", "Model", "Serial", "Status",
        "Qty", "Location", "Purchase", "Warranty", "Store",
        "Unit Price", "Depreciation (yrs)",
    ]
    rows_html = []
    for it in items:
        cells = [
            it.id, it.group_name, it.type, it.brand, it.model,
            it.serial, it.status, it.quantity, it.location,
            it.purchase_date, it.warranty_end, it.store,
            f"{it.unit_price:.2f}" if it.unit_price else "",
            f"{it.depreciation_years:.1f}" if it.depreciation_years else "",
        ]
        row = "".join(f"<td>{escape(str(c))}</td>" for c in cells)
        rows_html.append(f"<tr>{row}</tr>")

    head = "".join(f"<th>{h}</th>" for h in headers)
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>ThingKeeper Inventory</title>
<style>
  body {{ font-family: -apple-system, 'Segoe UI', sans-serif; margin: 24px; }}
  h1 {{ color: #305496; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 8px; text-align: left; }}
  th {{ background: #305496; color: #fff; position: sticky; top: 0; }}
  tr:nth-child(even) {{ background: #f7f7f7; }}
  .meta {{ color: #666; margin-bottom: 16px; }}
</style>
</head>
<body>
<h1>ThingKeeper Inventory</h1>
<p class="meta">Generated {datetime.now():%Y-%m-%d %H:%M} — {len(items)} items</p>
<table>
<thead><tr>{head}</tr></thead>
<tbody>
{chr(10).join(rows_html)}
</tbody>
</table>
</body>
</html>
"""
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(html)
    return path


def export_pdf_report(path: str | Path) -> Path:
    """Generate a PDF inventory report: summary + breakdown + warranty alerts."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    path = Path(path)
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("<b>ThingKeeper Inventory Report</b>", styles["Title"])
    story.append(title)
    story.append(Paragraph(
        f"Generated {datetime.now():%Y-%m-%d %H:%M}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 8 * mm))

    items = all_items()
    total_qty = total_quantity()
    tval = total_value()
    dval = total_depreciated_value()
    story.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
    summary = [
        ["Distinct items", str(len(items))],
        ["Total quantity", str(total_qty)],
        ["Total value (purchase)", f"{tval:,.2f}"],
        ["Total value (depreciated)", f"{dval:,.2f}"],
    ]
    for col, label in [("group_name", "By group"), ("status", "By status")]:
        summary.append([label, ", ".join(f"{k}={v}" for k, v in counts_by(col)) or "—"])
    t = Table(summary, colWidths=[55 * mm, 120 * mm])
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    # Group breakdown.
    story.append(Paragraph("<b>Breakdown by group</b>", styles["Heading2"]))
    grp_rows = [["Group", "Items", "Quantity"]]
    grp_counts = {k: v for k, v in counts_by("group_name")}
    qty_by_group: dict[str, int] = {}
    for it in items:
        g = it.group_name.strip() or "(none)"
        qty_by_group[g] = qty_by_group.get(g, 0) + it.quantity
    for g in sorted(qty_by_group):
        grp_rows.append([g, str(grp_counts.get(g, 0)), str(qty_by_group[g])])
    gt = Table(grp_rows, colWidths=[80 * mm, 40 * mm, 40 * mm])
    gt.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
    ]))
    story.append(gt)
    story.append(Spacer(1, 6 * mm))

    # Warranty alerts.
    expired = warranty_expired()
    expiring = warranty_expiring()
    story.append(Paragraph("<b>Warranty alerts</b>", styles["Heading2"]))
    story.append(Paragraph(
        f"Expired: {len(expired)} &nbsp;|&nbsp; "
        f"Expiring within {config.WARRANTY_SOON_DAYS} days: {len(expiring)}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 3 * mm))

    def _warranty_table(rows: list[Item], header_color: str):
        data = [["Warranty end", "Group", "Brand", "Model", "Serial"]]
        for it in rows:
            data.append([
                it.warranty_end, it.group_name, it.brand, it.model, it.serial,
            ])
        tbl = Table(data, colWidths=[28 * mm, 30 * mm, 35 * mm, 50 * mm, 30 * mm])
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        return tbl

    if expired:
        story.append(Paragraph("<b>Expired</b>", styles["Normal"]))
        story.append(_warranty_table(expired, "#963030"))
        story.append(Spacer(1, 3 * mm))
    if expiring:
        story.append(Paragraph("<b>Expiring soon</b>", styles["Normal"]))
        story.append(_warranty_table(expiring, "#967830"))

    doc.build(story)
    return path


# ----------------------------------------------------------- encrypted archives
_ENC_MAGIC = b"TKENC1"


def _derive_key(passphrase: str) -> bytes:
    import base64
    import hashlib

    digest = hashlib.sha256(passphrase.encode("utf-8")).digest()
    return base64.urlsafe_b64encode(digest)


def export_archive_encrypted(path: str | Path, passphrase: str) -> Path:
    """Write a passphrase-encrypted .tkz archive.

    Layout: 6-byte magic 'TKENC1' + Fernet token of the standard .tkz bytes.
    """
    from cryptography.fernet import Fernet

    if not passphrase:
        raise ValueError("Passphrase is required for encrypted archive")
    path = Path(path)
    plain = export_archive(path.with_suffix(".tkz.tmp"))
    try:
        key = _derive_key(passphrase)
        with open(plain, "rb") as f:
            data = f.read()
        token = Fernet(key).encrypt(data)
        with open(path, "wb") as f:
            f.write(_ENC_MAGIC + token)
    finally:
        plain.unlink(missing_ok=True)
    return path


def is_encrypted_archive(path: str | Path) -> bool:
    """Return True if the file is an encrypted .tkz archive."""
    try:
        with open(path, "rb") as f:
            return f.read(len(_ENC_MAGIC)) == _ENC_MAGIC
    except OSError:
        return False
